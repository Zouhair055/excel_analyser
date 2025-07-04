from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import os
import json
import glob
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from werkzeug.utils import secure_filename
import re

excel_bp = Blueprint('excel', __name__)

# Configuration
UPLOAD_FOLDER = os.path.abspath('uploads')
PROCESSED_FOLDER = os.path.abspath('processed')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def find_data_start_row(filepath):
    """🔍 Trouve la ligne où commencent vraiment les données"""
    print(f"🔍 Détection du début des données dans {filepath}")
    try:
        # Lire les 20 premières lignes sans en-têtes
        preview_df = pd.read_excel(filepath, header=None, nrows=20)
        
        for idx, row in preview_df.iterrows():
            # Chercher une ligne qui contient des en-têtes typiques
            row_str = ' '.join([str(val) for val in row if pd.notna(val)]).lower()
            
            # Mots-clés indiquant des en-têtes de données
            header_keywords = ['entity', 'date', 'transaction', 'period', 'amount', 'account', 'description', 'bank']
            
            if any(keyword in row_str for keyword in header_keywords):
                print(f"✅ Données détectées à partir de la ligne {idx + 1}")
                return idx
        
        # Si aucune ligne d'en-tête détectée, chercher la première ligne avec plusieurs valeurs non-nulles
        for idx, row in preview_df.iterrows():
            non_null_count = sum(1 for val in row if pd.notna(val) and str(val).strip() != '')
            if non_null_count >= 3:  # Au moins 3 colonnes avec des données
                print(f"✅ Données détectées à partir de la ligne {idx + 1} (par nombre de colonnes)")
                return idx
        
        print("⚠️ Impossible de détecter le début des données, utilisation de la ligne 1")
        return 0
        
    except Exception as e:
        print(f"⚠️ Erreur lors de la détection du début des données: {e}")
        return 0

def read_excel_smart(filepath):
    """📊 Lit le fichier Excel en détectant automatiquement où commencent les données"""
    print(f"📊 Lecture intelligente de {filepath}")
    
    start_row = find_data_start_row(filepath)
    
    # Lire le fichier en spécifiant la ligne de départ
    df = pd.read_excel(filepath, header=start_row)
    
    # Nettoyer les noms de colonnes (enlever les espaces, caractères bizarres)
    df.columns = [str(col).strip() if pd.notna(col) else f'Unnamed_{i}' for i, col in enumerate(df.columns)]
    
    # Supprimer les lignes complètement vides
    df = df.dropna(how='all')
    
    print(f"📊 Fichier lu avec succès: {df.shape[0]} lignes, {df.shape[1]} colonnes")
    print(f"📋 Colonnes détectées: {list(df.columns)}")
    
    return df

def clean_column_names(df):
    """🧹 Standardise les noms de colonnes avec plus de flexibilité"""
    print("🧹 Nettoyage des noms de colonnes...")
    
    # 🔍 DIAGNOSTIC : Avant nettoyage
    print(f"🔍 Colonnes AVANT nettoyage: {list(df.columns)}")
    
    column_mapping = {
        # Descriptions
        'description': 'Description', 'descrip': 'Descrip', 'desc': 'Description',
        'libelle': 'Description', 'libellé': 'Description', 'detail': 'Description',
     
        
        # Autres colonnes
        'nature': 'Nature', 'reference': 'Reference', 'service': 'Service', 
        'vessel': 'Vessel', 'amount': 'Amount CCYs', 'amount_usd': 'Amount USD', 
        'rate': 'Rate FX', 'bank': 'Bank account', 'ccy': 'CCY', 'currency': 'CCY'
    }
    
    rename_dict = {}
    for col in df.columns:
        if pd.isna(col) or col == '':
            continue
        
        col_lower = str(col).lower().strip()
        
        # Correspondance exacte
        if col_lower in column_mapping:
            rename_dict[col] = column_mapping[col_lower]
        else:
            # Correspondance partielle pour Description
            if any(desc_key in col_lower for desc_key in ['description', 'descrip', 'desc', 'libelle', 'detail']):
                if 'Description' not in df.columns:
                    rename_dict[col] = 'Description'
            
            # Autres correspondances partielles
            for old_name, new_name in column_mapping.items():
                if old_name in col_lower and new_name not in df.columns:
                    rename_dict[col] = new_name
                    break
    
    if rename_dict:
        df = df.rename(columns=rename_dict)
        print(f"🔄 Colonnes renommées: {rename_dict}")
    
    return df

def apply_formatting(df, filepath):
    """Applique un formatage Excel professionnel"""
    
    # 🔧 Formatter les colonnes numériques AVANT d'écrire dans Excel
    numeric_columns = ['Rate FX', 'Amount CCYs', 'Amount USD', 'Total']
    for col in numeric_columns:
        if col in df.columns:
            try:
                # Convertir en numérique, remplacer les erreurs par NaN
                df[col] = pd.to_numeric(df[col], errors='coerce')
                print(f"✅ Colonne '{col}' convertie en numérique")
            except Exception as e:
                print(f"⚠️ Erreur conversion numérique pour '{col}': {e}")
    
    wb = Workbook()
    ws = wb.active
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    data_font = Font(size=10)
    negative_font = Font(size=10, color="FF0000")
    
    # En-têtes
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Données avec formatage spécialisé
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            
            col_name = df.columns[col_idx - 1]
            
            # 💰 Format numérique pour les colonnes financières
            if col_name in numeric_columns:
                try:
                    if pd.notna(value) and value != '':
                        cell.value = float(value)
                        if col_name in ['Amount CCYs', 'Amount USD', 'Total']:
                            cell.number_format = '#,##0.00'  # Format monétaire
                        elif col_name == 'Rate FX':
                            cell.number_format = '0.00000'   # Format taux (5 décimales)
                        
                        # Couleur rouge pour les montants négatifs
                        if float(value) < 0:
                            cell.font = negative_font
                except (ValueError, TypeError):
                    pass
    
    # Ajuster largeurs et ajouter filtres
    for col_idx in range(1, len(df.columns) + 1):
        col_name = df.columns[col_idx - 1]
        if col_name in numeric_columns:
            ws.column_dimensions[get_column_letter(col_idx)].width = 18  # Plus large pour les nombres
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    
    wb.save(filepath)
    return True

def format_period_column(df):
        """📅 Formate la colonne Period au format 'mois-aa'"""
        if 'Period' not in df.columns:
            return df
        
        print("📅 Formatage de la colonne Period...")
        
        try:
            # Convertir en datetime si ce n'est pas déjà fait
            df['Period'] = pd.to_datetime(df['Period'], errors='coerce')
            
            # Créer le mapping français des mois
            mois_fr = {
                1: 'janv.', 2: 'févr.', 3: 'mars', 4: 'avr.',
                5: 'mai', 6: 'juin', 7: 'juil.', 8: 'août',
                9: 'sept.', 10: 'oct.', 11: 'nov.', 12: 'déc.'
            }
            
            # Formatter au format français
            def format_period(date):
                if pd.isna(date):
                    return ''
                try:
                    mois = mois_fr.get(date.month, str(date.month))
                    annee = str(date.year)[-2:]  # 2 derniers chiffres de l'année
                    return f"{mois}-{annee}"
                except:
                    return str(date)
            
            df['Period'] = df['Period'].apply(format_period)
            print("✅ Colonne Period formatée au format français")
            
        except Exception as e:
            print(f"⚠️ Erreur lors du formatage de Period: {e}")
        
        return df

class RulesPredictor:
    """🤖 Système de règles avec logs détaillés"""
    
    def __init__(self):
        self.rules = []
        self.stats = {'rules_loaded': 0, 'rules_applied': 0, 'cells_filled': 0}
    
    
    def load_rules(self):
        """📋 Charge les règles depuis les fichiers JSON"""
        print("📋 Chargement des règles...")
        
        # Chercher les fichiers de règles dans différents répertoires
        search_patterns = [
            "rules_*.json", 
            "**/rules_*.json", 
            "model_auto_remplissage/rules_*.json",
            "model_auto_remplissage/**/rules_*.json"
        ]
        
        rule_files = []
        for pattern in search_patterns:
            rule_files.extend(glob.glob(pattern, recursive=True))
        
        if not rule_files:
            print("❌ Aucun fichier de règles trouvé")
            return False
        
        # Prendre le fichier le plus récent
        latest_file = max(rule_files, key=os.path.getctime)
        print(f"📂 Fichier de règles sélectionné: {latest_file}")
        
        try:
            with open(latest_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.rules = data.get('rules', [])
            self.stats['rules_loaded'] = len(self.rules)
            print(f"✅ {len(self.rules)} règles chargées avec succès")
            
            # Afficher quelques exemples de règles
            if self.rules:
                print("📋 Exemples de règles chargées:")
                for i, rule in enumerate(self.rules[:3]):
                    pattern = rule.get('pattern', '')[:50]
                    cols = list(rule.get('fixed_columns', {}).keys())
                    print(f"   • Règle {i+1}: '{pattern}...' → Remplit {cols}")
            
            return True
            
        except Exception as e:
            print(f"❌ Erreur lors du chargement des règles: {e}")
            return False
    
    def apply_rules(self, df):
        """⚙️ Applique les règles avec logs détaillés"""
        print("⚙️ Application des règles de remplissage...")
        
        if not self.rules:
            print("⚠️ Aucune règle disponible")
            return df
        
        # Vérifier la présence de la colonne Description
        if 'Description' not in df.columns:
            print("❌ Colonne 'Description' manquante, impossible d'appliquer les règles")
            return df
        
        print(f"📊 Colonne Description détectée avec {df['Description'].notna().sum()} valeurs")
        
        # Colonnes protégées (ne jamais modifier)
        protected_columns = ['Amount CCYs', 'Amount USD', 'Rate FX', 'Entity', 'Period', 'Date']
        target_columns = ['Nature', 'Descrip', 'Vessel', 'Service', 'Reference']
        
        # Créer les colonnes cibles si elles n'existent pas
        for col in target_columns:
            if col not in df.columns:
                df[col] = ''
                print(f"➕ Colonne '{col}' créée")
        
        rules_applied = 0
        cells_filled = 0
        
        # 🎯 TRI OPTIMISÉ : Priorité + Support + Longueur du pattern
        def rule_priority(rule):
            priority = rule.get('priority', 999)  # Plus bas = meilleur
            support = rule.get('support', 0)     # Plus haut = meilleur
            pattern_len = len(rule.get('pattern', ''))  # Plus long = plus spécifique
            
            return (priority, -support, -pattern_len)  # Ordre croissant pour priority, décroissant pour autres
        
        sorted_rules = sorted(self.rules, key=rule_priority)
        print(f"🔄 Application de TOUTES les {len(sorted_rules)} règles dans l'ordre optimal...")
        
        for rule_idx, rule in enumerate(sorted_rules):  # ✅ TOUTES LES RÈGLES
            pattern = rule.get('pattern', '').lower().strip()
            if len(pattern) < 3:
                continue
            
            try:
                # 🔧 CORRECTION ROBUSTE : Gérer les DataFrame multiples
                if 'Description' not in df.columns:
                    if rule_idx < 5:  # Afficher seulement les 5 premières erreurs
                        print(f"   ⚠️ Règle {rule_idx+1}: Colonne 'Description' introuvable")
                    continue
                
                description_col = df['Description']
                
                # Forcer la conversion en Series si c'est un DataFrame
                if isinstance(description_col, pd.DataFrame):
                    description_series = description_col.iloc[:, 0]  # Prendre la première colonne
                    if rule_idx < 3:  # Afficher seulement les 3 premières conversions
                        print(f"   🔧 Règle {rule_idx+1}: DataFrame converti en Series")
                else:
                    description_series = description_col
                
                # Vérifier que c'est maintenant une Series
                if not isinstance(description_series, pd.Series):
                    if rule_idx < 5:  # Limiter les messages d'erreur
                        print(f"   ⚠️ Règle {rule_idx+1}: Type invalide après conversion (type: {type(description_series)})")
                    continue
                
                # Rechercher le pattern dans Description avec gestion d'erreur
                try:
                    mask = description_series.str.lower().str.contains(
                        re.escape(pattern), na=False, regex=True
                    )
                except Exception as str_error:
                    if rule_idx < 5:  # Limiter les messages d'erreur
                        print(f"   ⚠️ Règle {rule_idx+1}: Erreur de recherche pattern - {str_error}")
                    continue
                
                # Vérifier que mask est bien un boolean Series
                if not isinstance(mask, pd.Series) or mask.dtype != bool:
                    if rule_idx < 5:  # Limiter les messages d'erreur
                        print(f"   ⚠️ Règle {rule_idx+1}: Masque invalide")
                    continue
                
                matches_found = mask.sum()
                if matches_found == 0:
                    continue
                
                                # Afficher seulement un résumé compact
                if rule_idx < 3:  # Afficher seulement les 3 premières règles
                    print(f"🎯 Règle {rule_idx+1}: Pattern '{pattern[:30]}...' → {matches_found} correspondances")
                elif rule_idx == 3:
                    print(f"📝 Remplissage en cours... (mode silencieux)")
                
                # Appliquer les colonnes fixes SILENCIEUSEMENT
                rule_cells_filled = 0
                for col, value in rule.get('fixed_columns', {}).items():
                    if col in target_columns and col in df.columns:
                        try:
                            # Identifier les cellules vides dans les lignes correspondantes
                            empty_mask = (df[col].isna() | (df[col] == ''))
                            
                            # Combiner le masque de pattern avec le masque de cellules vides
                            final_mask = mask & empty_mask
                            
                            cells_to_fill = final_mask.sum()
                            if cells_to_fill > 0:
                                df.loc[final_mask, col] = value
                                rule_cells_filled += cells_to_fill
                                cells_filled += cells_to_fill
                                # 🔇 SILENCIEUX : Pas d'affichage détaillé
                        except Exception as col_error:
                            continue  # Mode silencieux pour les erreurs aussi
                
                if rule_cells_filled > 0:
                    rules_applied += 1
                            
            except Exception as e:
                if rule_idx < 5:  # Limiter les messages d'erreur
                    print(f"   ❌ Erreur règle {rule_idx+1}: {e}")
                continue
        
        # Afficher un résumé compact
        if len(sorted_rules) > 10:
            print(f"📊 Résumé: {rules_applied} règles actives sur {len(sorted_rules)} testées")
        
        self.stats.update({
            'rules_applied': rules_applied,
            'cells_filled': cells_filled
        })
        
        print(f"🎉 Remplissage terminé: {rules_applied}/{len(sorted_rules)} règles appliquées, {cells_filled} cellules remplies")
        
        return df

@excel_bp.route('/upload', methods=['POST'])
def upload_file():
    """📤 Upload et traitement du fichier Excel avec logs complets"""
    print("\n" + "="*60)
    print("🚀 DÉBUT DU TRAITEMENT EXCEL")
    print("="*60)
    
    try:
        if 'file' not in request.files:
            print("❌ Aucun fichier fourni dans la requête")
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '' or not allowed_file(file.filename):
            print(f"❌ Fichier invalide: {file.filename}")
            return jsonify({'error': 'Fichier invalide'}), 400
        
        print(f"📁 Fichier reçu: {file.filename}")
        
        # Sauvegarder le fichier
        filename = secure_filename(file.filename)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        print(f"💾 Fichier sauvegardé: {filepath}")
        
        # Lire et nettoyer les données avec détection intelligente
        df = read_excel_smart(filepath)
        df = clean_column_names(df)
        
        # 🎯 VALIDATION FLEXIBLE - Chercher une colonne similaire
        print("🔍 Recherche de la colonne Description...")
        description_col = None
        possible_desc_cols = ['Description', 'description', 'Descrip', 'descrip', 'Desc', 'desc', 
                              'Libelle', 'Libellé', 'Detail']  # ❌ RETIRE 'Transaction'
        
        # 🔧 CORRECTION : Chercher d'abord une vraie colonne Description
        for col in df.columns:
            col_lower = str(col).lower()
            if col_lower == 'description':  # Correspondance exacte d'abord
                description_col = col
                print(f"✅ Colonne Description exacte trouvée: '{col}'")
                break
        
        if description_col is None:
            for col in df.columns:
                if any(desc_name.lower() in str(col).lower() for desc_name in possible_desc_cols):
                    # 🔧 VÉRIFIER que ce n'est pas une colonne de date
                    if 'date' not in str(col).lower() and 'period' not in str(col).lower():
                        description_col = col
                        print(f"✅ Colonne description trouvée: '{col}'")
                        break
        
        if description_col is None:
            # Essayer de trouver une colonne avec du texte
            print("🔍 Recherche d'une colonne texte...")
            for col in df.columns:
                if df[col].dtype == 'object':  # Colonne texte
                    # 🔧 VÉRIFIER que ce n'est pas une colonne de date/période
                    if 'date' not in str(col).lower() and 'period' not in str(col).lower():
                        sample_values = df[col].dropna().head(3)
                        if len(sample_values) > 0 and all(len(str(val)) > 10 for val in sample_values):
                            description_col = col
                            print(f"✅ Colonne texte détectée: '{col}'")
                            break
        
        if description_col is None:
            available_cols = ', '.join(df.columns)
            print(f"❌ Aucune colonne Description trouvée. Colonnes disponibles: {available_cols}")
            return jsonify({'error': f'Aucune colonne Description trouvée. Colonnes disponibles: {available_cols}'}), 400
        
        # Renommer la colonne pour standardiser
        if description_col != 'Description':
            df = df.rename(columns={description_col: 'Description'})
            print(f"🔄 Colonne renommée: '{description_col}' → 'Description'")
        
        # Calculer qualité initiale
        initial_empty = df.isna().sum().sum() + (df == '').sum().sum()
        initial_completion = ((df.size - initial_empty) / df.size) * 100
        print(f"📊 Qualité initiale: {initial_completion:.1f}% rempli")
        
        # Appliquer les règles
        predictor = RulesPredictor()
        if predictor.load_rules():
            processed_df = predictor.apply_rules(df.copy())
        else:
            print("⚠️ Aucune règle chargée, fichier non modifié")
            processed_df = df.copy()
        # 📅 Formatter la colonne Period
        processed_df = format_period_column(processed_df)
        
        # Calculer qualité finale
        final_empty = processed_df.isna().sum().sum() + (processed_df == '').sum().sum()
        final_completion = ((processed_df.size - final_empty) / processed_df.size) * 100
        improvement = final_completion - initial_completion
        
        print(f"📊 Qualité finale: {final_completion:.1f}% rempli (+{improvement:.1f}%)")
        
        # Sauvegarder avec formatage
        processed_filename = f"processed_{unique_filename}"
        processed_filepath = os.path.join(PROCESSED_FOLDER, processed_filename)
        formatting_success = apply_formatting(processed_df, processed_filepath)
        
        print(f"💾 Fichier traité sauvegardé: {processed_filename}")
        print("🎉 TRAITEMENT TERMINÉ AVEC SUCCÈS")
        print("="*60)
        
        # 🔧 CONVERSION POUR JSON : Convertir les types numpy en types Python
        def convert_numpy_types(obj):
            """Convertit les types numpy en types Python pour JSON"""
            if isinstance(obj, (pd.Series, pd.DataFrame)):
                return obj.to_dict()
            elif hasattr(obj, 'item'):  # numpy types
                return obj.item()
            elif isinstance(obj, dict):
                return {key: convert_numpy_types(value) for key, value in obj.items()}
            elif isinstance(obj, list):
                return [convert_numpy_types(item) for item in obj]
            else:
                return obj
                
        return jsonify({
            'success': True,
            'message': 'Fichier traité avec succès',
            'processed_file': processed_filename,
            'original_file': filename,  # ✅ Ajout du nom original
            'columns_info': {
                'shape': [int(processed_df.shape[0]), int(processed_df.shape[1])],  # ✅ Format attendu
                'columns': list(processed_df.columns),
                'empty_columns': [col for col in processed_df.columns if processed_df[col].isna().all()]
            },
            'changes_applied': {
                'rules_applied': [f"Règle {i+1}" for i in range(min(5, int(predictor.stats['rules_applied'])))]  # ✅ Format attendu
            },
            'statistics': {
                'initial_completion': float(round(initial_completion, 2)),
                'final_completion': float(round(final_completion, 2)),
                'improvement': float(round(improvement, 2)),
                'cells_filled': int(predictor.stats['cells_filled']),
                'rules_applied': int(predictor.stats['rules_applied'])
            },
            'formatting_applied': formatting_success
        })
        
    except Exception as e:
        print(f"💥 ERREUR CRITIQUE: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500  # ✅ Ajout du return complet

@excel_bp.route('/download/<filename>')
def download_file(filename):
    """📥 Télécharge un fichier traité"""
    try:
        file_path = os.path.join(PROCESSED_FOLDER, filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'Fichier non trouvé'}), 404
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@excel_bp.route('/health')
def health_check():
    """🏥 Vérifie l'état du service"""
    predictor = RulesPredictor()
    rules_available = predictor.load_rules()
    
    return jsonify({
        'status': 'healthy',
        'rules_available': rules_available,
        'rules_count': len(predictor.rules) if rules_available else 0,
        'timestamp': datetime.datetime.now().isoformat()
    })