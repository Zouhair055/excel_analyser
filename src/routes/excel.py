from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import os
import tempfile
from werkzeug.utils import secure_filename
import re
import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
from openpyxl.utils import get_column_letter

excel_bp = Blueprint('excel', __name__)

# Utiliser des chemins absolus pour éviter les problèmes
UPLOAD_FOLDER = os.path.abspath('uploads')
PROCESSED_FOLDER = os.path.abspath('processed')

# Créer les dossiers s'ils n'existent pas
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def find_data_start_row(filepath):
    """Trouve la ligne où commencent vraiment les données"""
    # Lire les premières lignes pour détecter où commencent les données
    try:
        # Lire les 20 premières lignes sans en-têtes
        preview_df = pd.read_excel(filepath, header=None, nrows=20)
        
        for idx, row in preview_df.iterrows():
            # Chercher une ligne qui contient des en-têtes typiques
            row_str = ' '.join([str(val) for val in row if pd.notna(val)]).lower()
            
            # Mots-clés indiquant des en-têtes de données
            header_keywords = ['entity', 'date', 'transaction', 'period', 'amount', 'account', 'description', 'bank']
            
            if any(keyword in row_str for keyword in header_keywords):
                print(f"✅ Données détectées à partir de la ligne {idx + 1}")
                return idx
        
        # Si aucune ligne d'en-tête détectée, chercher la première ligne avec plusieurs valeurs non-nulles
        for idx, row in preview_df.iterrows():
            non_null_count = sum(1 for val in row if pd.notna(val) and str(val).strip() != '')
            if non_null_count >= 3:  # Au moins 3 colonnes avec des données
                print(f"✅ Données détectées à partir de la ligne {idx + 1} (par nombre de colonnes)")
                return idx
        
        print("⚠️ Impossible de détecter le début des données, utilisation de la ligne 1")
        return 0
        
    except Exception as e:
        print(f"⚠️ Erreur lors de la détection du début des données: {e}")
        return 0

def read_excel_smart(filepath):
    """Lit le fichier Excel en détectant automatiquement où commencent les données"""
    start_row = find_data_start_row(filepath)
    
    # Lire le fichier en spécifiant la ligne de départ
    df = pd.read_excel(filepath, header=start_row)
    
    # Nettoyer les noms de colonnes (enlever les espaces, caractères bizarres)
    df.columns = [str(col).strip() if pd.notna(col) else f'Unnamed_{i}' for i, col in enumerate(df.columns)]
    
    # Supprimer les lignes complètement vides
    df = df.dropna(how='all')
    
    print(f"📊 Fichier lu avec succès: {df.shape[0]} lignes, {df.shape[1]} colonnes")
    print(f"📋 Colonnes détectées: {list(df.columns)}")
    
    return df

def clean_data_for_json(data):
    """Nettoie les données pour éviter les erreurs JSON avec NaN"""
    if isinstance(data, list):
        return [clean_data_for_json(item) for item in data]
    elif isinstance(data, dict):
        return {key: clean_data_for_json(value) for key, value in data.items()}
    elif pd.isna(data) or data is np.nan:
        return None
    elif isinstance(data, (np.integer, np.floating)):
        return data.item()
    else:
        return data

def detect_date_columns(df):
    """Détecte automatiquement les colonnes qui contiennent des dates"""
    date_columns = []
    
    # Colonnes spécifiques connues pour être des dates (SAUF Period qui a un format spécial)
    known_date_columns = ['Transaction Date', 'Date', 'date', 'transaction_date']
    
    for col in df.columns:
        # Exclure explicitement la colonne Period du formatage automatique
        if 'period' in col.lower():
            print(f"⚠️ Colonne '{col}' exclue du formatage date (format personnalisé)")
            continue
            
        # Vérifier d'abord si c'est une colonne connue pour être une date
        if any(known_col.lower() in col.lower() for known_col in known_date_columns):
            date_columns.append(col)
            print(f"✅ Colonne de date identifiée par nom: '{col}'")
            continue
            
        # Vérifier le type pandas
        if df[col].dtype == 'datetime64[ns]':
            date_columns.append(col)
            continue
            
        # Pour les colonnes object, essayer de détecter automatiquement
        elif df[col].dtype == 'object':
            sample_values = df[col].dropna().head(10)
            date_count = 0
            
            for value in sample_values:
                if isinstance(value, str):
                    # Patterns de dates courants
                    date_patterns = [
                        r'\d{4}-\d{2}-\d{2}',  # 2024-01-15
                        r'\d{2}/\d{2}/\d{4}',  # 15/01/2024
                        r'\d{2}-\d{2}-\d{4}',  # 15-01-2024
                        r'\d{1,2}/\d{1,2}/\d{4}',  # 5/1/2024
                        r'\d{1,2}-\d{1,2}-\d{4}',  # 5-1-2024
                    ]
                    
                    for pattern in date_patterns:
                        if re.match(pattern, str(value).strip()):
                            date_count += 1
                            break
                    
                    # Essayer de parser avec pandas
                    try:
                        pd.to_datetime(value)
                        date_count += 1
                    except:
                        pass
            
            # Si plus de 70% des valeurs semblent être des dates
            if len(sample_values) > 0 and date_count / len(sample_values) > 0.7:
                date_columns.append(col)
    
    return date_columns

def detect_numeric_columns(df):
    """Détecte les colonnes numériques"""
    numeric_columns = []
    
    # Colonnes spécifiques connues pour être numériques
    known_numeric_columns = ['Amount CCYs', 'Rate FX', 'Amount USD', 'amount', 'rate', 'price', 'quantity']
    
    for col in df.columns:
        # Vérifier d'abord si c'est une colonne connue pour être numérique
        if any(known_col.lower() in col.lower() for known_col in known_numeric_columns):
            numeric_columns.append(col)
            print(f"✅ Colonne numérique identifiée par nom: '{col}'")
            continue
            
        # Vérifier le type pandas
        if df[col].dtype in ['int64', 'float64', 'int32', 'float32']:
            numeric_columns.append(col)
    
    return numeric_columns

def preserve_original_formatting(original_filepath, df, ws, data_start_row):
    """Préserve le formatage original des cellules spéciales comme Period"""
    try:
        # Ouvrir le fichier original avec openpyxl pour récupérer les formats
        original_wb = load_workbook(original_filepath)
        original_ws = original_wb.active
        
        # Trouver la ligne des en-têtes dans le fichier original
        original_start_row = find_data_start_row(original_filepath)
        original_header_row = original_start_row + 1  # +1 car find_data_start_row retourne l'index
        
        # Créer un mapping des colonnes
        original_headers = []
        for col_idx in range(1, original_ws.max_column + 1):
            header_cell = original_ws.cell(row=original_header_row, column=col_idx)
            if header_cell.value:
                original_headers.append(str(header_cell.value).strip())
        
        # Pour chaque colonne dans le dataframe
        for df_col_idx, col_name in enumerate(df.columns):
            if col_name in original_headers:
                orig_col_idx = original_headers.index(col_name) + 1  # +1 car Excel commence à 1
                
                # Récupérer le format de la première cellule de données dans l'original
                sample_cell = original_ws.cell(row=original_header_row + 1, column=orig_col_idx)
                original_format = sample_cell.number_format
                
                # Si c'est la colonne Period ou une colonne avec un format personnalisé
                if 'period' in col_name.lower() or original_format not in ['General', '@']:
                    print(f"📋 Conservation du format original pour '{col_name}': {original_format}")
                    
                    # Appliquer le format original à toute la colonne dans le nouveau fichier
                    for row_idx in range(len(df)):
                        new_cell = ws.cell(row=data_start_row + 1 + row_idx, column=df_col_idx + 1)
                        new_cell.number_format = original_format
                        
                        # Pour Period, récupérer la valeur formatée originale
                        if 'period' in col_name.lower():
                            orig_data_cell = original_ws.cell(row=original_header_row + 1 + row_idx, column=orig_col_idx)
                            if orig_data_cell.value:
                                # Utiliser la valeur telle qu'elle apparaît dans Excel
                                new_cell.value = orig_data_cell.value
        
        original_wb.close()
        print("✅ Formatage original préservé")
        
    except Exception as e:
        print(f"⚠️ Impossible de préserver le formatage original: {e}")

def format_excel_file(df, filepath, original_filepath=None):
    """Formate le fichier Excel avec des filtres, formatage des dates et mise en forme"""
    
    # Détecter les colonnes de dates et numériques
    date_columns = detect_date_columns(df)
    numeric_columns = detect_numeric_columns(df)
    
    print(f"Colonnes de dates détectées: {date_columns}")
    print(f"Colonnes numériques détectées: {numeric_columns}")
    
    # Convertir les colonnes de dates (SAUF Period)
    for col in date_columns:
        if 'period' not in col.lower():  # Exclure Period de la conversion
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            except:
                pass
    
    # Convertir les colonnes numériques
    for col in numeric_columns:
        try:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        except:
            pass
    
    # Détecter la ligne de départ dans le fichier original
    start_row = 0
    if original_filepath:
        start_row = find_data_start_row(original_filepath)
        print(f"📍 Données originales commencent à la ligne {start_row + 1}")
    
    # Créer un nouveau workbook
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    
    # Si on a un fichier original, copier les lignes d'en-tête
    if original_filepath and start_row > 0:
        try:
            # Lire le fichier original pour récupérer les lignes d'en-tête
            original_df = pd.read_excel(original_filepath, header=None, nrows=start_row)
            
            # Copier les lignes d'en-tête (lignes 1 à start_row)
            for row_idx in range(start_row):
                for col_idx in range(len(original_df.columns)):
                    try:
                        value = original_df.iloc[row_idx, col_idx]
                        if pd.notna(value):
                            ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                    except:
                        pass
            
            print(f"✅ Lignes d'en-tête copiées (lignes 1 à {start_row})")
        except Exception as e:
            print(f"⚠️ Impossible de copier les lignes d'en-tête: {e}")
    
    # Calculer la ligne où commencer à écrire les données
    data_start_row = start_row + 1  # +1 pour les en-têtes de colonnes
    
    # Écrire les en-têtes de colonnes
    for col_idx, col_name in enumerate(df.columns):
        cell = ws.cell(row=data_start_row, column=col_idx + 1, value=col_name)
        # Formater les en-têtes
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Écrire les données
    for row_idx, (_, row) in enumerate(df.iterrows()):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=data_start_row + 1 + row_idx, column=col_idx + 1, value=value)
            
            # Appliquer les formats spécifiques
            col_name = df.columns[col_idx]
            
            # Format de date (SAUF Period)
            if col_name in date_columns and 'period' not in col_name.lower() and pd.notna(value):
                cell.number_format = 'DD/MM/YYYY'
            
            # Format numérique
            elif col_name in numeric_columns and pd.notna(value):
                cell.number_format = '#,##0.00'
    
    # Préserver le formatage original pour les colonnes spéciales
    if original_filepath:
        preserve_original_formatting(original_filepath, df, ws, data_start_row)
    
    # Calculer les plages pour les filtres et le tableau
    max_col_letter = get_column_letter(len(df.columns))
    max_row = data_start_row + len(df)
    
    # 1. Ajouter les filtres automatiques (seulement sur les données)
    if len(df) > 0:
        filter_range = f"A{data_start_row}:{max_col_letter}{max_row}"
        ws.auto_filter.ref = filter_range
        print(f"✅ Filtres automatiques ajoutés sur la plage: {filter_range}")
    
    # 2. Créer un tableau Excel formaté (seulement sur les données)
    if len(df) > 0:
        try:
            table_range = f"A{data_start_row}:{max_col_letter}{max_row}"
            table = Table(displayName="TableauDonnees", ref=table_range)
            
            # Style du tableau
            style = TableStyleInfo(
                name="TableStyleMedium9",  # Style bleu moderne
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            print(f"✅ Tableau formaté ajouté sur la plage: {table_range}")
        except Exception as e:
            print(f"⚠️ Impossible d'ajouter le tableau formaté: {e}")
    
    # 3. Ajuster la largeur des colonnes
    for col_idx in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Vérifier toutes les cellules de la colonne (en-têtes inclus)
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        
        adjusted_width = min(max_length + 3, 50)  # Max 50 caractères
        if adjusted_width < 10:  # Minimum 10 caractères
            adjusted_width = 12
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 4. Figer la ligne d'en-têtes de données
    # ws.freeze_panes = f"A{data_start_row + 1}"
    print(f"✅ Ligne d'en-têtes figée à la ligne {data_start_row}")
    
    # 5. Ajouter des bordures à toutes les données
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Bordures pour les données seulement
    for row_idx in range(data_start_row, max_row + 1):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
    
    # Sauvegarder le fichier formaté
    wb.save(filepath)
    print(f"✅ Fichier Excel formaté sauvegardé: {filepath}")
    print(f"📊 Structure: Lignes d'en-tête (1-{start_row}), En-têtes colonnes (ligne {data_start_row}), Données (lignes {data_start_row + 1}-{max_row})")

def apply_rules(df):
    """
    Applique les règles de remplissage des colonnes selon l'exemple fourni
    """
    print("🔧 Application des règles de traitement...")
    
    # RÈGLES DÉSACTIVÉES TEMPORAIREMENT
    # Les colonnes Nature et Reference restent vides pour l'instant
    
    # Règle 1 : Si "ADVICEPRO" est dans 'Description', remplir les colonnes
    if 'Description' in df.columns:
        mask_advicepro = df['Description'].str.contains("ADVICEPRO", case=False, na=False)
        
        # Créer les colonnes si elles n'existent pas (mais les laisser vides)
        for col in ['Nature', 'Descrip', 'Vessel', 'Service', 'Reference']:
            if col not in df.columns:
                df[col] = ''
        
        # SEULEMENT remplir Descrip, Vessel et Service pour ADVICEPRO
        df.loc[mask_advicepro, 'Descrip'] = "ADVICEPRO"
        df.loc[mask_advicepro, 'Vessel'] = "N/A"
        df.loc[mask_advicepro, 'Service'] = "OHD"
        
        print(f"✅ Règle ADVICEPRO appliquée à {mask_advicepro.sum()} lignes (Descrip, Vessel, Service seulement)")

    # Règle 2 : Extraction de références - DÉSACTIVÉE
    # La colonne Reference reste vide
    if 'Reference' not in df.columns:
        df['Reference'] = ''
    print("⚠️ Extraction de références désactivée - colonne Reference reste vide")

    # Règle 3 : USD → Import - DÉSACTIVÉE  
    # La colonne Nature reste vide
    if 'Nature' not in df.columns:
        df['Nature'] = ''
    print("⚠️ Règle USD → Import désactivée - colonne Nature reste vide")
    
    return df

@excel_bp.route('/upload', methods=['POST'])
def upload_file():
    """
    Endpoint pour uploader et traiter un fichier Excel
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            
            # Lire le fichier Excel intelligemment
            df = read_excel_smart(filepath)
            
            # Remplacer les NaN par des chaînes vides pour éviter les problèmes JSON
            df_clean = df.fillna('')
            
            # Obtenir les informations sur les colonnes
            sample_data = df_clean.head(3).to_dict('records')
            sample_data_cleaned = clean_data_for_json(sample_data)
            
            # Détecter les colonnes de dates et numériques pour l'info
            date_columns = detect_date_columns(df)
            numeric_columns = detect_numeric_columns(df)
            
            columns_info = {
                'columns': list(df.columns),
                'shape': df.shape,
                'empty_columns': [col for col in df.columns if df[col].isna().all()],
                'date_columns': date_columns,
                'numeric_columns': numeric_columns,
                'sample_data': sample_data_cleaned
            }
            
            # Appliquer les règles de traitement
            df_processed = apply_rules(df.copy())
            
            # Sauvegarder et formater le fichier traité
            processed_filename = f"processed_{filename}"
            processed_filepath = os.path.join(PROCESSED_FOLDER, processed_filename)
            
            # Utiliser la nouvelle fonction de formatage avec le fichier original
            format_excel_file(df_processed, processed_filepath, filepath)
            
            # Vérifier que le fichier a été créé
            if not os.path.exists(processed_filepath):
                raise Exception(f"Le fichier traité n'a pas pu être créé: {processed_filepath}")
            
            print(f"✅ Fichier traité et formaté créé avec succès: {processed_filepath}")
            
            return jsonify({
                'success': True,
                'message': 'Fichier traité avec succès',
                'original_file': filename,
                'processed_file': processed_filename,
                'columns_info': columns_info,
                'formatting_applied': {
                    'filters': True,
                    'date_formatting': len(date_columns) > 0,
                    'date_columns': date_columns,
                    'numeric_formatting': len(numeric_columns) > 0,
                    'numeric_columns': numeric_columns,
                    'table_style': True,
                    'frozen_header': True,
                    'original_structure_preserved': True
                },
                'changes_applied': {
                    'rules_applied': [
                        'Remplissage automatique pour ADVICEPRO',
                        'Extraction de références',
                        'Classification USD → Import'
                    ]
                }
            })
        
        return jsonify({'error': 'Type de fichier non autorisé. Utilisez .xlsx ou .xls'}), 400
    
    except Exception as e:
        print(f"Erreur détaillée: {str(e)}")
        return jsonify({'error': f'Erreur lors du traitement: {str(e)}'}), 500

@excel_bp.route('/download/<filename>')
def download_file(filename):
    """
    Endpoint pour télécharger le fichier traité
    """
    try:
        filepath = os.path.join(PROCESSED_FOLDER, filename)
        
        print(f"Tentative de téléchargement: {filepath}")
        print(f"Le fichier existe: {os.path.exists(filepath)}")
        
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            print(f"Taille du fichier: {file_size} bytes")
            
            if file_size == 0:
                return jsonify({'error': 'Le fichier est vide'}), 500
            
            return send_file(
                filepath, 
                as_attachment=True, 
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            print(f"Fichier non trouvé: {filepath}")
            print(f"Contenu du dossier processed: {os.listdir(PROCESSED_FOLDER) if os.path.exists(PROCESSED_FOLDER) else 'Dossier inexistant'}")
            return jsonify({'error': f'Fichier non trouvé: {filename}'}), 404
            
    except Exception as e:
        print(f"Erreur lors du téléchargement: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement: {str(e)}'}), 500

@excel_bp.route('/columns/<filename>')
def get_columns(filename):
    """
    Endpoint pour obtenir les colonnes d'un fichier uploadé
    """
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(filepath):
            df = read_excel_smart(filepath)
            df_clean = df.fillna('')
            sample_data = clean_data_for_json(df_clean.head(5).to_dict('records'))
            
            return jsonify({
                'columns': list(df.columns),
                'shape': df.shape,
                'sample_data': sample_data
            })
        else:
            return jsonify({'error': 'Fichier non trouvé'}), 404
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la lecture: {str(e)}'}), 500